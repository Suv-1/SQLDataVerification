from openpyxl import Workbook
import mysql.connector
from datetime import datetime
import pandas as pd
from flask import Flask, render_template, request
import webbrowser
import os 
import win32com.client

app = Flask(__name__, template_folder="template")

@app.route("/submit", methods= ["POST"])
def getValue():
    id,password,host,database, file = "","","","",""
    
    id = request.form.get("user")
    password = request.form.get("password")
    host = request.form.get("host")
    database = request.form.get("database")
    file = request.form.get("file")
    mail = request.form.get("email")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    df = pd.read_excel(file)
    output_file = f"C:\\Users\\subhankar.samal\\Desktop\\\\Output Folder\\output_file_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Table Summary"
    ws["A1"] = "Table Name"
    ws["B1"] = "Condition"
    ws["C1"] = "Availability"
    ws["D1"] = "Accessibility"
    ws["E1"] = "Row Count"


    cn = mysql.connector.connect(host=host,database=database,user=id, password=password)
    cursor = cn.cursor()
        

    for index, row in df.iterrows():
        table_name = row["Table"]
        condition = row["Condition"]
        cursor.execute(f"Show tables like '{table_name}'")
        result = cursor.fetchone()
        if not result:
            ws.cell(row= index+2, column = 1, value = table_name)
            ws.cell(row= index+2, column = 2, value = condition)
            ws.cell(row= index+2, column = 3, value = "Not Available")
            ws.cell(row= index+2, column = 4, value = "-")
            ws.cell(row= index+2, column = 5, value = "-")
            continue

        cursor.execute(f"select count(*) from {table_name}")
        try: cursor.fetchone()
        except mysql.connector.errors.ProgrammingError:
            ws.cell(row= index+2, column = 1, value = table_name)
            ws.cell(row= index+2, column = 2, value = condition)
            ws.cell(row= index+2, column = 3, value = "Available")
            ws.cell(row= index+2, column = 4, value = "Not Accessible")
            ws.cell(row= index+2, column = 5, value = "-")
            continue

        cursor.execute(f"select count(*) from {table_name} {condition}")
        row_count = cursor.fetchone()[0]
        ws.cell(row= index+2, column = 1, value = table_name)
        ws.cell(row= index+2, column = 2, value = condition)
        ws.cell(row= index+2, column = 3, value = "Available")
        ws.cell(row= index+2, column = 4, value = "Accessible")
        ws.cell(row= index+2, column = 5, value = row_count)

    wb.save(output_file)

    cursor.close()
    cn.close()
    if len(mail)!=0:
        sendEmail(mail,"Table Load Validation","Output Table",output_file)
        message = f"Output file {output_file} has been sent to {mail}"
        return f"<script>alert('{message}'); window.location.href='/';</script>"
    message = f"Output File Created: {output_file}"
    return f"<script>alert('{message}'); window.location.href='/';</script>"



def sendEmail(recipient, subject, body, attachment):
    ol = win32com.client.Dispatch('Outlook.Application')
    email = ol.CreateItem(0)
    email.To = recipient
    email.Subject = subject
    email.Body = body
    attach = os.path.abspath(attachment)
    email.Attachments.Add(attach)
    email.Send()
@app.route('/')
def index():
    return render_template('index1.html')

if __name__=="__main__":
    webbrowser.open_new("http://127.0.0.1:5000")
    app.run(debug=True)
